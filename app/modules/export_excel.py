"""Excel export with clickable links."""

from typing import List
import pandas as pd
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

from app.schemas import SupervisorProfile


def clean_name_for_export(name: str) -> str:
    """Clean name by removing titles and fixing 'essor' issue before export."""
    if not name:
        return ""
    
    # Remove extra whitespace
    name = " ".join(name.split())
    
    # Remove common prefixes (case-insensitive, anywhere in the string)
    prefixes = ["Dr.", "Dr", "Prof.", "Prof", "Professor", "Mr.", "Mr", "Mrs.", "Mrs", "Ms.", "Ms"]
    for prefix in prefixes:
        # Remove from start
        if name.lower().startswith(prefix.lower() + " "):
            name = name[len(prefix):].strip()
        elif name.lower().startswith(prefix.lower()):
            name = name[len(prefix):].strip()
        # Also remove if it appears anywhere
        pattern = r'\b' + re.escape(prefix) + r'\b'
        name = re.sub(pattern, '', name, flags=re.IGNORECASE).strip()
    
    # Remove "essor" if it appears at the start (leftover from "Professor")
    if name.lower().startswith("essor "):
        name = name[6:].strip()
    elif name.lower().startswith("essor"):
        name = name[5:].strip()
    
    # Clean up multiple spaces
    name = " ".join(name.split())
    
    return name.strip()


def export_to_excel(profiles: List[SupervisorProfile], output_path: Path) -> None:
    """Export supervisor profiles to Excel with clickable links."""
    
    # Convert to records - only include specified fields
    records = []
    for p in profiles:
        # Clean name to ensure no "Professor" or "essor" in Excel output
        cleaned_name = clean_name_for_export(p.name)
        
        records.append({
            "Institution": p.institution,
            "QS Rank": p.qs_rank,
            "Region": p.region,
            "Country": p.country,
            "Title": p.title,
            "Name": cleaned_name,
            "Email": p.email,
            "Profile URL": p.profile_url,
            "Keywords": ", ".join(p.keywords) if p.keywords else "",
            "Fit Score": round(p.fit_score, 3),  # 3 decimal places
            "Tier": p.tier
        })
    
    df = pd.DataFrame(records)
    
    # Create workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Supervisors"
    
    # Add headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(df.columns, 1):
            value = record[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Make URLs clickable
            if col_name == "Profile URL" and value:
                if isinstance(value, str) and (value.startswith("http://") or value.startswith("https://")):
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")
    
    # Adjust column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 15)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
    
    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Exported {len(profiles)} supervisors to {output_path}")

